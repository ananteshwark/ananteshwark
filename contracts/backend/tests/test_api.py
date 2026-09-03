"""API integration tests against the real app (see conftest.py)."""
import uuid


def _new_email():
    return f"user-{uuid.uuid4().hex[:8]}@example.com"


class TestUserManagement:
    def test_create_update_deactivate_delete(self, client, admin_headers):
        email = _new_email()
        # create
        r = client.post("/api/auth/users", headers=admin_headers,
                        json={"email": email, "name": "Temp", "password": "password123", "role": "VIEWER"})
        assert r.status_code == 200, r.text
        uid = r.json()["id"]
        assert r.json()["role"] == "VIEWER"

        # promote to validator
        r = client.patch(f"/api/auth/users/{uid}", headers=admin_headers, json={"role": "VALIDATOR"})
        assert r.status_code == 200 and r.json()["role"] == "VALIDATOR"

        # deactivate -> that user can no longer log in
        r = client.patch(f"/api/auth/users/{uid}", headers=admin_headers, json={"is_active": False})
        assert r.status_code == 200 and r.json()["is_active"] is False
        r = client.post("/api/auth/login", json={"email": email, "password": "password123"})
        assert r.status_code == 401

        # reactivate + reset password, then login with the new password
        client.patch(f"/api/auth/users/{uid}", headers=admin_headers, json={"is_active": True})
        r = client.post(f"/api/auth/users/{uid}/reset-password", headers=admin_headers,
                        json={"new_password": "brandnew123"})
        assert r.status_code == 200
        r = client.post("/api/auth/login", json={"email": email, "password": "brandnew123"})
        assert r.status_code == 200

        # soft delete -> gone from list and cannot log in
        r = client.delete(f"/api/auth/users/{uid}", headers=admin_headers)
        assert r.status_code == 200
        listed = client.get("/api/auth/users", headers=admin_headers).json()
        assert uid not in [u["id"] for u in listed]
        r = client.post("/api/auth/login", json={"email": email, "password": "brandnew123"})
        assert r.status_code == 401

    def test_cannot_delete_or_demote_last_admin(self, client, admin_headers):
        me = client.get("/api/auth/me", headers=admin_headers).json()
        # deleting yourself is blocked
        r = client.delete(f"/api/auth/users/{me['id']}", headers=admin_headers)
        assert r.status_code == 400
        # demoting the only admin is blocked
        r = client.patch(f"/api/auth/users/{me['id']}", headers=admin_headers, json={"role": "VIEWER"})
        assert r.status_code == 400

    def test_change_own_password(self, client, admin_headers):
        email = _new_email()
        client.post("/api/auth/users", headers=admin_headers,
                    json={"email": email, "name": "Pw", "password": "initial123", "role": "VIEWER"})
        token = client.post("/api/auth/login", json={"email": email, "password": "initial123"}).json()["token"]
        h = {"Authorization": f"Bearer {token}"}
        # wrong current password rejected
        r = client.post("/api/auth/change-password", headers=h,
                        json={"current_password": "wrong", "new_password": "updated123"})
        assert r.status_code == 401
        # correct current password succeeds
        r = client.post("/api/auth/change-password", headers=h,
                        json={"current_password": "initial123", "new_password": "updated123"})
        assert r.status_code == 200
        assert client.post("/api/auth/login", json={"email": email, "password": "updated123"}).status_code == 200

    def test_role_enforcement(self, client, admin_headers):
        # a viewer cannot access admin-only endpoints
        email = _new_email()
        client.post("/api/auth/users", headers=admin_headers,
                    json={"email": email, "name": "V", "password": "viewer12345", "role": "VIEWER"})
        token = client.post("/api/auth/login", json={"email": email, "password": "viewer12345"}).json()["token"]
        h = {"Authorization": f"Bearer {token}"}
        assert client.get("/api/auth/users", headers=h).status_code == 403
        assert client.get("/api/settings", headers=h).status_code == 403
        assert client.get("/api/audit", headers=h).status_code == 403
        # but read-only screens are fine
        assert client.get("/api/dashboard", headers=h).status_code == 200


class TestContractGroup:
    def _mk(self, db, **kw):
        from app.models import Contract, ContractStatus
        c = Contract(status=ContractStatus.VALIDATED, raw_extracted={}, confidence={}, **kw)
        db.add(c); db.flush()
        return c

    def test_link_groups_documents_and_shows_renewal_history(self, client, admin_headers):
        import datetime
        from app.database import SessionLocal
        db = SessionLocal()
        # An MSA with a renewal, plus a separate NDA — all one logical contract.
        msa_v1 = self._mk(db, vendor_name_raw="Acme", contract_type="MSA",
                          start_date=datetime.date(2023, 1, 1), end_date=datetime.date(2023, 12, 31))
        msa_v1.thread_id = msa_v1.sr_no
        msa_v2 = self._mk(db, vendor_name_raw="Acme", contract_type="MSA",
                          start_date=datetime.date(2024, 1, 1), end_date=datetime.date(2024, 12, 31),
                          renews_contract_id=msa_v1.sr_no, thread_id=msa_v1.sr_no)
        nda = self._mk(db, vendor_name_raw="Acme", contract_type="NDA",
                       start_date=datetime.date(2023, 1, 1), end_date=datetime.date(2025, 12, 31))
        db.commit()
        msa_sr, msa2_sr, nda_sr = msa_v1.sr_no, msa_v2.sr_no, nda.sr_no
        db.close()

        # link the NDA into the MSA's group
        r = client.post(f"/api/contracts/{msa_sr}/link-document", headers=admin_headers,
                        json={"sr_no": nda_sr})
        assert r.status_code == 200, r.text
        g = r.json()
        member_srs = {m["sr_no"] for m in g["members"]}
        assert {msa_sr, nda_sr} <= member_srs
        # combined renewal history includes the MSA renewal (msa_v2) even though it
        # was only linked via the MSA's thread — ordered by start date (oldest first)
        hist_srs = [h["sr_no"] for h in g["renewal_history"]]
        assert hist_srs == [msa_sr, nda_sr, msa2_sr]
        # unlink removes the NDA from the group
        assert client.post(f"/api/contracts/{nda_sr}/unlink-document", headers=admin_headers).status_code == 200
        g2 = client.get(f"/api/contracts/{msa_sr}/group", headers=admin_headers).json()
        assert nda_sr not in {m["sr_no"] for m in g2["members"]}


class TestExpiringFilter:
    def test_expiring_days_filters_by_window(self, client, admin_headers):
        import datetime
        import uuid
        from app.database import SessionLocal
        from app.models import Contract, ContractStatus
        today = datetime.date.today()
        tag = uuid.uuid4().hex[:8]
        db = SessionLocal()
        db.add(Contract(vendor_name_raw=f"soon-{tag}", status=ContractStatus.VALIDATED,
                        end_date=today + datetime.timedelta(days=20), raw_extracted={}, confidence={}))
        db.add(Contract(vendor_name_raw=f"later-{tag}", status=ContractStatus.VALIDATED,
                        end_date=today + datetime.timedelta(days=200), raw_extracted={}, confidence={}))
        db.commit(); db.close()

        r = client.get(f"/api/contracts?q={tag}&expiring_days=30", headers=admin_headers).json()
        names = [c["vendor_name"] for c in r["items"]]
        assert any(n.startswith("soon-") for n in names)
        assert not any(n.startswith("later-") for n in names)


class TestDashboardActiveValue:
    def test_department_value_counts_only_active_contracts(self, client, admin_headers):
        import datetime
        import uuid
        from app.database import SessionLocal
        from app.models import Contract, ContractStatus, Department
        today = datetime.date.today()
        db = SessionLocal()
        dept = Department(name=f"ActiveVal-{uuid.uuid4().hex[:6]}")
        db.add(dept); db.flush()
        did = dept.id
        # active validated (future end) -> counts
        db.add(Contract(vendor_name_raw="A", department_id=did, status=ContractStatus.VALIDATED,
                        end_date=today + datetime.timedelta(days=30), contract_value=100,
                        raw_extracted={}, confidence={}))
        # validated but expired (past end) -> excluded
        db.add(Contract(vendor_name_raw="B", department_id=did, status=ContractStatus.VALIDATED,
                        end_date=today - datetime.timedelta(days=5), contract_value=1000,
                        raw_extracted={}, confidence={}))
        # pending (not validated) -> excluded
        db.add(Contract(vendor_name_raw="C", department_id=did, status=ContractStatus.PENDING_VALIDATION,
                        end_date=today + datetime.timedelta(days=30), contract_value=500,
                        raw_extracted={}, confidence={}))
        db.commit(); db.close()

        data = client.get("/api/dashboard", headers=admin_headers).json()
        row = next(d for d in data["departments"] if d["name"] == dept.name)
        assert row["contract_count"] == 1
        assert row["total_value"] == 100.0


class TestSigningEntityFilter:
    def test_filter_and_distinct_list_and_dashboard_widget(self, client, admin_headers):
        import datetime
        import uuid
        from app.database import SessionLocal
        from app.models import Contract, ContractStatus
        today = datetime.date.today()
        ent_a = f"EntityA-{uuid.uuid4().hex[:6]}"
        ent_b = f"EntityB-{uuid.uuid4().hex[:6]}"
        db = SessionLocal()
        # Two active validated contracts under ent_a, one under ent_b.
        db.add(Contract(vendor_name_raw="a1", signing_entity=ent_a, status=ContractStatus.VALIDATED,
                        end_date=today + datetime.timedelta(days=30), contract_value=100,
                        raw_extracted={}, confidence={}))
        db.add(Contract(vendor_name_raw="a2", signing_entity=ent_a, status=ContractStatus.VALIDATED,
                        end_date=today + datetime.timedelta(days=30), contract_value=50,
                        raw_extracted={}, confidence={}))
        db.add(Contract(vendor_name_raw="b1", signing_entity=ent_b, status=ContractStatus.VALIDATED,
                        end_date=today + datetime.timedelta(days=30), contract_value=200,
                        raw_extracted={}, confidence={}))
        db.commit(); db.close()

        # Distinct entities endpoint includes both
        entities = client.get("/api/contracts/signing-entities", headers=admin_headers).json()["entities"]
        assert ent_a in entities and ent_b in entities

        # Filtering by signing_entity narrows the list
        r = client.get(f"/api/contracts?signing_entity={ent_a}", headers=admin_headers).json()
        assert r["total"] == 2
        assert all(c["signing_entity"] == ent_a for c in r["items"])

        # Dashboard entity widget aggregates count + value
        data = client.get("/api/dashboard", headers=admin_headers).json()
        row = next(e for e in data["entities"] if e["name"] == ent_a)
        assert row["contract_count"] == 2
        assert row["total_value"] == 150.0


class TestFieldSuggestions:
    def test_learns_department_from_vendor_history(self, client, admin_headers):
        import uuid
        from app.database import SessionLocal
        from app.models import Contract, ContractStatus, Department, Vendor
        from app.services.vendor_matching import normalize_vendor_name
        db = SessionLocal()
        vname = f"LearnVendor-{uuid.uuid4().hex[:6]}"
        vendor = Vendor(name=vname, normalized_name=normalize_vendor_name(vname))
        dept = Department(name=f"LearnDept-{uuid.uuid4().hex[:6]}")
        db.add_all([vendor, dept]); db.flush()
        vid, did = vendor.id, dept.id
        # Two validated contracts for this vendor, both in the same department.
        for _ in range(2):
            db.add(Contract(vendor_id=vid, department_id=did, signing_entity="TruBridge",
                            status=ContractStatus.VALIDATED, raw_extracted={}, confidence={}))
        # A pending contract for the same vendor with no department yet.
        pending = Contract(vendor_id=vid, status=ContractStatus.PENDING_VALIDATION,
                           raw_extracted={}, confidence={})
        db.add(pending); db.commit(); sr = pending.sr_no; db.close()

        r = client.get(f"/api/contracts/{sr}/field-suggestions", headers=admin_headers)
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["history_count"] == 2
        dept_s = next(s for s in body["suggestions"] if s["field"] == "department_id")
        assert dept_s["suggested"] == did
        assert dept_s["suggested_label"] == dept.name
        assert dept_s["current_empty"] is True


class TestAudit:
    def test_global_audit_records_user_actions(self, client, admin_headers):
        # creating a user writes an audit row
        client.post("/api/auth/users", headers=admin_headers,
                    json={"email": _new_email(), "name": "Aud", "password": "password123", "role": "VIEWER"})
        r = client.get("/api/audit?entity_type=user", headers=admin_headers)
        assert r.status_code == 200
        data = r.json()
        assert data["total"] >= 1
        assert any(item["action"] == "CREATE" for item in data["items"])


class TestContractsPaginationAndRestore:
    def _seed_contract(self, status="PENDING_VALIDATION"):
        from app.database import SessionLocal
        from app.models import Contract, ContractStatus
        db = SessionLocal()
        c = Contract(vendor_name_raw="PagerVendor", contract_service="svc",
                     status=ContractStatus(status), raw_extracted={}, confidence={})
        db.add(c); db.commit(); sr = c.sr_no; db.close()
        return sr

    def test_pagination_envelope(self, client, admin_headers):
        for _ in range(3):
            self._seed_contract()
        r = client.get("/api/contracts?limit=2&offset=0", headers=admin_headers)
        assert r.status_code == 200
        body = r.json()
        assert "total" in body and body["total"] >= 3
        assert len(body["items"]) == 2

    def test_export_filtered_view_returns_xlsx(self, client, admin_headers):
        import io
        from openpyxl import load_workbook
        self._seed_contract()  # PENDING_VALIDATION
        # unique vendor so we can assert the filter narrows the export
        from app.database import SessionLocal
        from app.models import Contract, ContractStatus
        db = SessionLocal()
        c = Contract(vendor_name_raw="ExportOnlyVendor", contract_service="svc",
                     status=ContractStatus("VALIDATED"), raw_extracted={}, confidence={})
        db.add(c); db.commit(); db.close()

        r = client.get("/api/contracts/export?q=ExportOnlyVendor", headers=admin_headers)
        assert r.status_code == 200, r.text
        assert r.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        body = "\n".join(
            " ".join(str(cell) for cell in row if cell is not None)
            for row in ws.iter_rows(values_only=True)
        )
        assert "ExportOnlyVendor" in body
        assert "PagerVendor" not in body  # the q filter excluded it

    def test_reject_then_restore(self, client, admin_headers):
        sr = self._seed_contract()
        r = client.post(f"/api/contracts/{sr}/reject", headers=admin_headers, json={"reason": "not a contract"})
        assert r.status_code == 200 and r.json()["status"] == "REJECTED"
        # restore brings it back to the queue
        r = client.post(f"/api/contracts/{sr}/restore", headers=admin_headers)
        assert r.status_code == 200 and r.json()["status"] == "PENDING_VALIDATION"
        assert r.json()["rejection_reason"] is None
        # a fresh pending contract cannot be restored
        sr2 = self._seed_contract()
        assert client.post(f"/api/contracts/{sr2}/restore", headers=admin_headers).status_code == 400


class TestEmailTest:
    def test_dry_run_email_test(self, client, admin_headers):
        # default env has EMAIL_DRY_RUN=true, so this logs instead of sending
        r = client.post("/api/settings/email-test", headers=admin_headers, json={"to": "someone@example.com"})
        assert r.status_code == 200, r.text
        assert r.json()["dry_run"] is True


class TestVendorMerge:
    def test_merge_repoints_contracts_and_folds_aliases(self, client, admin_headers):
        # two vendors, each with a contract
        target = client.post("/api/vendors", headers=admin_headers,
                             json={"name": "Acme Technologies Pvt Ltd", "aliases": ["Acme"]}).json()
        source = client.post("/api/vendors", headers=admin_headers,
                             json={"name": "ACME Tech LLP", "aliases": ["ACME Tech"], "addresses": ["1 St"]}).json()

        from app.database import SessionLocal
        from app.models import Contract, ContractStatus
        db = SessionLocal()
        c = Contract(vendor_id=source["id"], contract_service="svc",
                     status=ContractStatus.VALIDATED, raw_extracted={}, confidence={})
        db.add(c); db.commit(); csr = c.sr_no; db.close()

        r = client.post(f"/api/vendors/{target['id']}/merge", headers=admin_headers,
                        json={"source_ids": [source["id"]]})
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["contracts_moved"] == 1
        assert "ACME Tech LLP" in body["absorbed"]

        # the moved contract now points at the target vendor
        moved = client.get(f"/api/contracts/{csr}", headers=admin_headers).json()
        assert moved["vendor_id"] == target["id"]
        # source is gone from the vendor list, its name became a target alias
        vendors = client.get("/api/vendors", headers=admin_headers).json()
        assert source["id"] not in [v["id"] for v in vendors]
        tgt = next(v for v in vendors if v["id"] == target["id"])
        assert any("ACME Tech LLP" == a or "ACME Tech" == a for a in tgt["aliases"])

    def test_cannot_merge_into_missing_target(self, client, admin_headers):
        r = client.post("/api/vendors/999999/merge", headers=admin_headers, json={"source_ids": [1]})
        assert r.status_code == 404

    def test_merge_is_reversible(self, client, admin_headers):
        target = client.post("/api/vendors", headers=admin_headers,
                             json={"name": f"Keep-{uuid.uuid4().hex[:6]} Pvt Ltd", "addresses": ["HQ"]}).json()
        source = client.post("/api/vendors", headers=admin_headers,
                             json={"name": f"Fold-{uuid.uuid4().hex[:6]} LLP", "aliases": ["FoldAlias"],
                                   "addresses": ["Branch St"]}).json()
        from app.database import SessionLocal
        from app.models import Contract, ContractStatus
        db = SessionLocal()
        c = Contract(vendor_id=source["id"], contract_service="svc",
                     status=ContractStatus.VALIDATED, raw_extracted={}, confidence={})
        db.add(c); db.commit(); csr = c.sr_no; db.close()

        # merge, capturing the batch id
        res = client.post(f"/api/vendors/{target['id']}/merge", headers=admin_headers,
                          json={"source_ids": [source["id"]]}).json()
        batch = res["batch_id"]
        assert batch

        # the merge shows up as reversible
        merges = client.get("/api/vendors/merges?active_only=true", headers=admin_headers).json()
        assert any(m["batch_id"] == batch for m in merges)

        # undo it
        r = client.post(f"/api/vendors/merges/{batch}/undo", headers=admin_headers)
        assert r.status_code == 200, r.text
        assert r.json()["count"] == 1

        # contract is back on the source vendor
        moved = client.get(f"/api/contracts/{csr}", headers=admin_headers).json()
        assert moved["vendor_id"] == source["id"]
        # source vendor is active again
        vendors = client.get("/api/vendors", headers=admin_headers).json()
        assert source["id"] in [v["id"] for v in vendors]
        # the folded alias/address were removed from the survivor
        tgt = client.get(f"/api/vendors/{target['id']}/history", headers=admin_headers).json()["vendor"]
        assert "FoldAlias" not in tgt["aliases"]
        assert "Branch St" not in (tgt["addresses"] or [])
        # merge no longer in the active list; undoing again is a 404
        merges = client.get("/api/vendors/merges?active_only=true", headers=admin_headers).json()
        assert not any(m["batch_id"] == batch for m in merges)
        assert client.post(f"/api/vendors/merges/{batch}/undo", headers=admin_headers).status_code == 404


class TestDepartments:
    def test_update_and_soft_delete(self, client, admin_headers):
        created = client.post("/api/departments", headers=admin_headers,
                              json={"name": f"Dept-{uuid.uuid4().hex[:6]}"}).json()
        did = created["id"]
        # rename + set a default recipient
        r = client.put(f"/api/departments/{did}", headers=admin_headers,
                       json={"name": "Procurement", "default_recipient_email": "proc@example.com"})
        assert r.status_code == 200, r.text
        assert r.json()["name"] == "Procurement"
        assert r.json()["default_recipient_email"] == "proc@example.com"
        # reflected in the master list
        listed = client.get("/api/departments", headers=admin_headers).json()
        assert any(d["id"] == did and d["name"] == "Procurement" for d in listed)
        # soft delete removes it from the list
        assert client.delete(f"/api/departments/{did}", headers=admin_headers).status_code == 200
        listed2 = client.get("/api/departments", headers=admin_headers).json()
        assert not any(d["id"] == did for d in listed2)
        # updating a deleted department 404s
        assert client.put(f"/api/departments/{did}", headers=admin_headers,
                          json={"name": "X"}).status_code == 404

    def test_multiple_recipient_emails(self, client, admin_headers):
        r = client.post("/api/departments", headers=admin_headers, json={
            "name": f"Multi-{uuid.uuid4().hex[:6]}",
            "default_recipient_email": "a@example.com, b@example.com\n c@example.com",
        })
        assert r.status_code == 200, r.text
        # normalized to a comma-separated list
        assert r.json()["default_recipient_email"] == "a@example.com, b@example.com, c@example.com"

    def test_rejects_invalid_recipient_email(self, client, admin_headers):
        r = client.post("/api/departments", headers=admin_headers, json={
            "name": f"Bad-{uuid.uuid4().hex[:6]}",
            "default_recipient_email": "a@example.com, not-an-email",
        })
        assert r.status_code == 422

    def test_department_recipients_resolve_for_reminders(self, client, admin_headers):
        from app.database import SessionLocal
        from app.models import Contract, ContractStatus, Department
        from app.services.reminders import resolve_recipients
        db = SessionLocal()
        dept = Department(name=f"Recip-{uuid.uuid4().hex[:6]}",
                          default_recipient_email="one@example.com, two@example.com")
        db.add(dept); db.flush()
        c = Contract(vendor_name_raw="RV", status=ContractStatus.VALIDATED,
                     department_id=dept.id, raw_extracted={}, confidence={})
        db.add(c); db.commit()
        emails, primary = resolve_recipients(db, c)
        db.close()
        assert emails == ["one@example.com", "two@example.com"]
        assert primary == "one@example.com"


class TestManualContract:
    def test_create_and_validate(self, client, admin_headers):
        dept = client.post("/api/departments", headers=admin_headers, json={"name": f"Dept-{uuid.uuid4().hex[:6]}"}).json()
        r = client.post("/api/contracts", headers=admin_headers, json={
            "signing_entity": "TruBridge", "new_vendor_name": "Manual Vendor Co",
            "start_date": "2025-01-01", "end_date": "2025-12-31",
            "department_id": dept["id"], "contract_service": "Manual service",
            "po_number": "PO-MANUAL-1",
        })
        assert r.status_code == 200, r.text
        c = r.json()
        assert c["status"] == "PENDING_VALIDATION"
        assert c["extraction_model"] == "manual-entry"
        assert c["vendor_name"] == "Manual Vendor Co"
        # it validates (all mandatory fields present)
        sr = c["sr_no"]
        v = client.post(f"/api/contracts/{sr}/validate", headers=admin_headers, json={"force": True})
        assert v.status_code == 200 and v.json()["status"] == "VALIDATED"

    def test_upload_document(self, client, admin_headers):
        c = client.post("/api/contracts", headers=admin_headers,
                        json={"signing_entity": "Arai", "contract_service": "x"}).json()
        sr = c["sr_no"]
        files = {"file": ("contract.pdf", b"%PDF-1.4 test", "application/pdf")}
        r = client.post(f"/api/contracts/{sr}/upload", headers=admin_headers, files=files)
        assert r.status_code == 200, r.text
        assert r.json()["contract_link"].endswith("contract.pdf")


class TestMandatoryPO:
    def test_po_number_required_for_validation(self, client, admin_headers):
        dept = client.post("/api/departments", headers=admin_headers,
                           json={"name": f"PO-{uuid.uuid4().hex[:6]}"}).json()
        c = client.post("/api/contracts", headers=admin_headers, json={
            "signing_entity": "TruBridge", "new_vendor_name": "PO Vendor Co",
            "start_date": "2025-01-01", "end_date": "2025-12-31",
            "department_id": dept["id"], "contract_service": "svc",
        }).json()
        sr = c["sr_no"]
        # no PO -> blocked
        r = client.post(f"/api/contracts/{sr}/validate", headers=admin_headers, json={"force": True})
        assert r.status_code == 422
        assert "PO Number" in r.json()["detail"]
        # with PO -> validates
        r2 = client.post(f"/api/contracts/{sr}/validate", headers=admin_headers,
                         json={"force": True, "po_number": "PO-1"})
        assert r2.status_code == 200 and r2.json()["status"] == "VALIDATED"


class TestIngestionConfidence:
    def _seed(self, status="PENDING_VALIDATION", confidence=None):
        from app.database import SessionLocal
        from app.models import Contract, ContractStatus, IngestionFile, IngestionStatus
        db = SessionLocal()
        c = Contract(vendor_name_raw="ConfVendor", status=ContractStatus[status],
                     raw_extracted={}, confidence=confidence or {})
        db.add(c); db.flush()
        f = IngestionFile(path="/tmp/x.pdf", filename=f"conf-{uuid.uuid4().hex[:6]}.pdf",
                          sha256=uuid.uuid4().hex, status=IngestionStatus.PENDING_VALIDATION,
                          contract_id=c.sr_no)
        db.add(f); db.commit()
        fid, sr, name = f.id, c.sr_no, f.filename
        db.close()
        return fid, sr, name

    def test_min_confidence_and_low_flag_in_list(self, client, admin_headers):
        fid, sr, name = self._seed(confidence={"vendor": 0.9, "start_date": 0.4})
        items = client.get(f"/api/ingestion?q={name}", headers=admin_headers).json()["items"]
        row = next(r for r in items if r["id"] == fid)
        assert row["min_confidence"] == 0.4
        assert row["low_confidence"] is True

    def test_retry_low_confidence_supersedes_pending_contract(self, client, admin_headers):
        from app.database import SessionLocal
        from app.models import Contract
        fid, sr, name = self._seed(confidence={"vendor": 0.3})
        r = client.post(f"/api/ingestion/{fid}/retry", headers=admin_headers)
        assert r.status_code == 200, r.text
        assert r.json()["status"] == "QUEUED" and r.json()["contract_id"] is None
        db = SessionLocal()
        superseded = db.get(Contract, sr)
        assert superseded.deleted_at is not None
        db.close()

    def test_cannot_retry_after_validation(self, client, admin_headers):
        fid, sr, name = self._seed(status="VALIDATED", confidence={"vendor": 0.3})
        r = client.post(f"/api/ingestion/{fid}/retry", headers=admin_headers)
        assert r.status_code == 400

    def test_completeness_present_in_list(self, client, admin_headers):
        fid, sr, name = self._seed(confidence={"vendor": 0.9})
        body = client.get(f"/api/ingestion?q={name}", headers=admin_headers).json()
        assert "secondary_enabled" in body
        row = next(r for r in body["items"] if r["id"] == fid)
        assert row["completeness"] is not None

    def test_secondary_retry_requires_enabled(self, client, admin_headers):
        fid, sr, name = self._seed(confidence={"vendor": 0.3})
        # disabled by default -> 400
        client.put("/api/settings", headers=admin_headers,
                   json={"values": {"secondary_extraction_enabled": "false"}})
        r = client.post(f"/api/ingestion/{fid}/retry?secondary=true", headers=admin_headers)
        assert r.status_code == 400
        # enable -> allowed, re-queues
        client.put("/api/settings", headers=admin_headers,
                   json={"values": {"secondary_extraction_enabled": "true"}})
        fid2, sr2, name2 = self._seed(confidence={"vendor": 0.3})
        r2 = client.post(f"/api/ingestion/{fid2}/retry?secondary=true", headers=admin_headers)
        assert r2.status_code == 200 and r2.json()["status"] == "QUEUED"
        client.put("/api/settings", headers=admin_headers,
                   json={"values": {"secondary_extraction_enabled": "false"}})


class TestVendorAutoAdd:
    def test_auto_create_and_dedupe_from_extraction(self, client, admin_headers):
        from app.database import SessionLocal
        from app.models import Vendor
        from app.services.extraction_worker import _resolve_or_create_vendor
        db = SessionLocal()
        vid1 = _resolve_or_create_vendor(db, "Acme Global Pvt Ltd")
        db.commit()
        assert vid1 is not None
        # a suffix/case variation normalizes to the same vendor — no duplicate
        vid2 = _resolve_or_create_vendor(db, "ACME GLOBAL")
        db.commit()
        assert vid2 == vid1
        assert db.query(Vendor).filter(Vendor.normalized_name == "acme global").count() == 1
        assert _resolve_or_create_vendor(db, "   ") is None
        db.close()


class TestAITags:
    def test_apply_ai_tags_matches_and_creates(self, client, admin_headers):
        from app.database import SessionLocal
        from app.models import Contract, ContractStatus, Tag
        from app.services.extraction_worker import _apply_ai_tags
        db = SessionLocal()
        db.add(Tag(name="Legal")); db.commit()
        c = Contract(vendor_name_raw="TagVendor", status=ContractStatus.PENDING_VALIDATION,
                     raw_extracted={}, confidence={})
        db.add(c); db.flush()
        _apply_ai_tags(db, c, ["legal", "high-value", "legal"])  # match existing (ci) + create + dup
        db.commit()
        names = sorted(t.name for t in c.tags)
        assert names == ["Legal", "high-value"]
        # the existing "Legal" tag was reused, not duplicated
        assert db.query(Tag).filter(Tag.name == "Legal").count() == 1
        db.close()


class TestContractImport:
    CSV = (
        "Sr No,Signing Entity,Vendor,Start Date,End Date,Department,Contract Service,Contract Value,PO Number\n"
        "1,TruBridge,Imported Vendor A,01/01/2025,31/12/2025,ImportDept,Support,50000,PO-A\n"   # complete -> validated
        "2,TruBridge,Imported Vendor B,2025-02-01,2026-01-31,ImportDept,Audit,,PO-B\n"
        "3,,,badstuff,,,,,\n"                                                                   # errors
    )

    def test_dry_run_then_import(self, client, admin_headers):
        files = {"file": ("register.csv", self.CSV.encode(), "text/csv")}
        # dry run persists nothing
        r = client.post("/api/contracts/import?dry_run=true", headers=admin_headers, files=files)
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["total_rows"] == 3
        assert body["created"] == 2  # third row has errors
        assert len(body["row_errors"]) == 1
        assert "ImportDept" in body["created_departments"]

        # nothing was actually created
        before = client.get("/api/contracts?q=Imported Vendor A", headers=admin_headers).json()["total"]
        assert before == 0

        # real import
        files = {"file": ("register.csv", self.CSV.encode(), "text/csv")}
        r = client.post("/api/contracts/import?dry_run=false", headers=admin_headers, files=files)
        assert r.status_code == 200
        assert r.json()["created"] == 2
        after = client.get("/api/contracts?q=Imported Vendor A", headers=admin_headers).json()
        assert after["total"] == 1
        # complete row imported as VALIDATED
        assert after["items"][0]["status"] == "VALIDATED"

    def test_rejects_unknown_columns(self, client, admin_headers):
        files = {"file": ("x.csv", b"Foo,Bar\n1,2\n", "text/csv")}
        r = client.post("/api/contracts/import", headers=admin_headers, files=files)
        assert r.status_code == 400


class TestAttachments:
    def _seed(self):
        from app.database import SessionLocal
        from app.models import Contract, ContractStatus
        db = SessionLocal()
        c = Contract(vendor_name_raw="AttachVendor", status=ContractStatus.VALIDATED,
                     raw_extracted={}, confidence={})
        db.add(c); db.commit(); sr = c.sr_no; db.close()
        return sr

    def test_add_list_download_delete(self, client, admin_headers):
        sr = self._seed()
        files = {"file": ("amendment1.pdf", b"%PDF-1.4 amendment", "application/pdf")}
        r = client.post(f"/api/contracts/{sr}/attachments?kind=amendment", headers=admin_headers, files=files)
        assert r.status_code == 200, r.text
        aid = r.json()["id"]
        assert r.json()["kind"] == "amendment"

        listed = client.get(f"/api/contracts/{sr}/attachments", headers=admin_headers).json()
        assert len(listed) == 1 and listed[0]["filename"] == "amendment1.pdf"

        dl = client.get(f"/api/contracts/{sr}/attachments/{aid}/file", headers=admin_headers)
        assert dl.status_code == 200 and b"amendment" in dl.content

        assert client.delete(f"/api/contracts/{sr}/attachments/{aid}", headers=admin_headers).status_code == 200
        assert client.get(f"/api/contracts/{sr}/attachments", headers=admin_headers).json() == []


class TestLoginLockout:
    def test_lockout_after_repeated_failures(self, client):
        email = f"lockme-{uuid.uuid4().hex[:8]}@example.com"
        # 5 wrong-password attempts (default LOGIN_MAX_ATTEMPTS=5)
        codes = [
            client.post("/api/auth/login", json={"email": email, "password": "wrong"}).status_code
            for _ in range(5)
        ]
        assert all(c == 401 for c in codes)
        # the next attempt is throttled
        r = client.post("/api/auth/login", json={"email": email, "password": "wrong"})
        assert r.status_code == 429
        assert "Retry-After" in r.headers


class TestFullTextSearch:
    def _seed_with_text(self, text):
        from app.database import SessionLocal
        from app.models import Contract, ContractStatus
        db = SessionLocal()
        c = Contract(vendor_name_raw="TextVendor", contract_service="svc",
                     status=ContractStatus.VALIDATED, raw_extracted={}, confidence={},
                     extracted_text=text)
        db.add(c); db.commit(); sr = c.sr_no; db.close()
        return sr

    def test_global_search_returns_text_matches_with_snippet(self, client, admin_headers):
        token = f"zorptoken{uuid.uuid4().hex[:6]}"
        self._seed_with_text(f"This agreement contains a special {token} clause deep inside the body.")
        r = client.get(f"/api/search?q={token}", headers=admin_headers)
        assert r.status_code == 200
        matches = r.json()["text_matches"]
        assert len(matches) >= 1
        assert any(token in (m["snippet"] or "") for m in matches)

    def test_contract_list_in_text_flag(self, client, admin_headers):
        token = f"bodyterm{uuid.uuid4().hex[:6]}"
        self._seed_with_text(f"Nothing structured, but the word {token} appears in the document text only.")
        # without in_text, the token isn't in any structured field -> no hit
        assert client.get(f"/api/contracts?q={token}", headers=admin_headers).json()["total"] == 0
        # with in_text, it matches the document body
        assert client.get(f"/api/contracts?q={token}&in_text=true", headers=admin_headers).json()["total"] == 1


class TestDashboardCharts:
    def test_expiry_by_month_present(self, client, admin_headers):
        d = client.get("/api/dashboard", headers=admin_headers).json()
        assert "expiry_by_month" in d
        assert len(d["expiry_by_month"]) == 12
        assert all("month" in m and "count" in m for m in d["expiry_by_month"])


class TestContractNotes:
    def _seed(self):
        from app.database import SessionLocal
        from app.models import Contract, ContractStatus
        db = SessionLocal()
        c = Contract(vendor_name_raw="NoteVendor", status=ContractStatus.VALIDATED,
                     raw_extracted={}, confidence={})
        db.add(c); db.commit(); sr = c.sr_no; db.close()
        return sr

    def test_add_list_delete(self, client, admin_headers):
        sr = self._seed()
        r = client.post(f"/api/contracts/{sr}/notes", headers=admin_headers, json={"body": "Check the indemnity clause"})
        assert r.status_code == 200, r.text
        nid = r.json()["id"]
        listed = client.get(f"/api/contracts/{sr}/notes", headers=admin_headers).json()
        assert len(listed) == 1 and listed[0]["body"] == "Check the indemnity clause"
        assert listed[0]["author"]
        assert client.delete(f"/api/contracts/{sr}/notes/{nid}", headers=admin_headers).status_code == 200
        assert client.get(f"/api/contracts/{sr}/notes", headers=admin_headers).json() == []

    def test_empty_note_rejected(self, client, admin_headers):
        sr = self._seed()
        assert client.post(f"/api/contracts/{sr}/notes", headers=admin_headers, json={"body": ""}).status_code == 422

    def test_viewer_can_read_but_not_write(self, client, admin_headers):
        sr = self._seed()
        email = _new_email()
        client.post("/api/auth/users", headers=admin_headers,
                    json={"email": email, "name": "V", "password": "viewer12345", "role": "VIEWER"})
        token = client.post("/api/auth/login", json={"email": email, "password": "viewer12345"}).json()["token"]
        h = {"Authorization": f"Bearer {token}"}
        assert client.get(f"/api/contracts/{sr}/notes", headers=h).status_code == 200
        assert client.post(f"/api/contracts/{sr}/notes", headers=h, json={"body": "x"}).status_code == 403


class TestFailureAlerts:
    def test_email_recipients_and_enabled_flag(self, client):
        from app.database import SessionLocal
        from app.models import IngestionFile
        from app.services import failure_alerts, notifications
        from app.services.settings_store import set_setting

        db = SessionLocal()
        set_setting(db, "failure_alerts_enabled", "true")
        set_setting(db, "failure_alert_emails", "ops@example.com")
        set_setting(db, "failure_alert_webhook", "")
        db.commit()
        rec = IngestionFile(path="/x/bad.pdf", filename="bad.pdf", sha256="hh")
        db.add(rec); db.commit()

        sent = []
        original = notifications.CHANNELS["email"].send
        notifications.CHANNELS["email"].send = lambda to, subject, body, cc=None: sent.append((to, subject))
        try:
            failure_alerts.notify_extraction_failure(db, rec, "boom")
            assert sent and "ops@example.com" in sent[0][0]
            assert "bad.pdf" in sent[0][1]

            # disabled -> no alert
            sent.clear()
            set_setting(db, "failure_alerts_enabled", "false"); db.commit()
            failure_alerts.notify_extraction_failure(db, rec, "boom")
            assert sent == []
        finally:
            notifications.CHANNELS["email"].send = original
            set_setting(db, "failure_alerts_enabled", "true")
            set_setting(db, "failure_alert_emails", "")
            db.commit()
            db.close()


class TestReminderScheduleAndRenewal:
    def _seed_validated(self, **kw):
        from datetime import date, timedelta
        from app.database import SessionLocal
        from app.models import Contract, ContractStatus
        db = SessionLocal()
        defaults = dict(
            signing_entity="TruBridge", vendor_name_raw="RenewVendor",
            contract_service="Support", contract_tenure="1 year",
            end_date=date.today() + timedelta(days=40),
            status=ContractStatus.VALIDATED, raw_extracted={}, confidence={},
            custom_offsets=[30, 7],
        )
        defaults.update(kw)
        c = Contract(**defaults)
        db.add(c); db.commit(); sr = c.sr_no; db.close()
        return sr

    def test_reminder_schedule_preview(self, client, admin_headers):
        sr = self._seed_validated()
        r = client.get(f"/api/contracts/{sr}/reminder-schedule", headers=admin_headers)
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["stopped"] is False
        assert len(body["dates"]) >= 1  # at least the 30-days-before date is upcoming

    def test_schedule_stopped_when_terminated(self, client, admin_headers):
        from app.models import LifecycleStatus
        sr = self._seed_validated(lifecycle_status=LifecycleStatus.TERMINATED)
        body = client.get(f"/api/contracts/{sr}/reminder-schedule", headers=admin_headers).json()
        assert body["stopped"] is True
        assert body["stopped_reason"] == "TERMINATED"
        assert body["dates"] == []

    def test_renew_creates_linked_draft(self, client, admin_headers):
        from datetime import date, timedelta
        sr = self._seed_validated()
        # attach a recipient so we can check it's copied
        client.put(f"/api/contracts/{sr}/recipients", headers=admin_headers,
                   json={"recipients": [{"name": "Owner", "email": "owner@example.com", "is_primary": True}]})
        src = client.get(f"/api/contracts/{sr}", headers=admin_headers).json()

        r = client.post(f"/api/contracts/{sr}/renew", headers=admin_headers)
        assert r.status_code == 200, r.text
        renewal = r.json()
        assert renewal["status"] == "PENDING_VALIDATION"
        assert renewal["renews_contract_id"] == sr
        assert renewal["thread_id"] in (sr, src.get("thread_id") or sr)
        assert renewal["contract_service"] == "Support"
        # start date = source end date + 1 day
        assert renewal["start_date"] == (date.fromisoformat(src["end_date"]) + timedelta(days=1)).isoformat()
        # tenure carried forward -> end date derived
        assert renewal["end_date"] is not None
        # recipient copied
        assert any(rr["email"] == "owner@example.com" for rr in renewal["recipients"])

    def test_cannot_renew_pending(self, client, admin_headers):
        from app.database import SessionLocal
        from app.models import Contract, ContractStatus
        db = SessionLocal()
        c = Contract(vendor_name_raw="X", status=ContractStatus.PENDING_VALIDATION, raw_extracted={}, confidence={})
        db.add(c); db.commit(); sr = c.sr_no; db.close()
        assert client.post(f"/api/contracts/{sr}/renew", headers=admin_headers).status_code == 400


class TestHealth:
    def test_liveness(self, client):
        assert client.get("/api/health").json()["status"] == "ok"

    def test_readiness_checks_db(self, client):
        r = client.get("/api/health/ready")
        assert r.status_code == 200
        assert r.json()["database"] == "ok"


class TestBulkActions:
    def _seed(self, n, status="PENDING_VALIDATION"):
        from app.database import SessionLocal
        from app.models import Contract, ContractStatus
        db = SessionLocal()
        srs = []
        for _ in range(n):
            c = Contract(vendor_name_raw="BulkVendor", contract_service="svc",
                         status=ContractStatus(status), raw_extracted={}, confidence={})
            db.add(c); db.flush(); srs.append(c.sr_no)
        db.commit(); db.close()
        return srs

    def test_bulk_assign_department(self, client, admin_headers):
        srs = self._seed(3)
        dept = client.post("/api/departments", headers=admin_headers, json={"name": f"Bulk-{uuid.uuid4().hex[:6]}"}).json()
        r = client.post("/api/contracts/bulk", headers=admin_headers,
                        json={"sr_nos": srs, "action": "assign_department", "department_id": dept["id"]})
        assert r.status_code == 200, r.text
        assert r.json()["updated_count"] == 3
        got = client.get(f"/api/contracts/{srs[0]}", headers=admin_headers).json()
        assert got["department_id"] == dept["id"]

    def test_bulk_reject(self, client, admin_headers):
        srs = self._seed(2)
        r = client.post("/api/contracts/bulk", headers=admin_headers,
                        json={"sr_nos": srs, "action": "reject", "reason": "bulk cleanup"})
        assert r.status_code == 200 and r.json()["updated_count"] == 2
        assert client.get(f"/api/contracts/{srs[0]}", headers=admin_headers).json()["status"] == "REJECTED"

    def test_bulk_validate_skips_incomplete(self, client, admin_headers):
        # seeded contracts lack mandatory fields -> all skipped, none validated
        srs = self._seed(2)
        r = client.post("/api/contracts/bulk", headers=admin_headers,
                        json={"sr_nos": srs, "action": "validate"})
        assert r.status_code == 200
        body = r.json()
        assert body["updated_count"] == 0
        assert len(body["skipped"]) == 2
        assert all("missing" in s["reason"] for s in body["skipped"])

    def test_bulk_unknown_action(self, client, admin_headers):
        r = client.post("/api/contracts/bulk", headers=admin_headers,
                        json={"sr_nos": [1], "action": "nonsense"})
        assert r.status_code == 400


def test_mandatory_fields_are_published_for_the_form(client, admin_headers):
    """The form marks required fields from this list rather than its own copy —
    a second list is how a form ends up disagreeing with what the server rejects."""
    from app.api.contracts_api import MANDATORY_FIELDS
    out = client.get("/api/contracts/mandatory-fields", headers=admin_headers).json()["mandatory"]
    assert {m["field"] for m in out} == {f for f, _ in MANDATORY_FIELDS}
    assert all(m["label"] and m["form_field"] for m in out)
    # The vendor is picked by name in the form but stored as an id.
    assert next(m["form_field"] for m in out if m["field"] == "vendor") == "vendor_name_raw"
