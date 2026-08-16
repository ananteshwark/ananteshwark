"""Excel exports in the organization's contract register format."""
import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Font

REGISTER_COLUMNS = [
    "Sr No",
    "Signing Entity",
    "Vendor",
    "Vendor Address",
    "Start Date",
    "End Date",
    "Contract Tenure",
    "Department",
    "PO Number",
    "Contract Value",
    "IKS Signing Authority",
    "Vendor Signing Authority",
    "Contract Service",
    "Service Summary",
    "Payment Term",
    "Notice Period",
    "Contract Link",
]


def _register_row(c) -> list:
    return [
        c.sr_no,
        c.signing_entity,
        c.vendor.name if c.vendor else c.vendor_name_raw,
        c.vendor_address,
        c.start_date.isoformat() if c.start_date else None,
        c.end_date.isoformat() if c.end_date else None,
        c.contract_tenure,
        c.department.name if c.department else None,
        c.po_number,
        float(c.contract_value) if c.contract_value is not None else None,
        c.iks_signing_authority,
        c.vendor_signing_authority,
        c.contract_service,
        c.service_summary,
        c.payment_term,
        c.notice_period,
        c.contract_link,
    ]


def contracts_to_register_csv(contracts) -> bytes:
    """The same register columns as the XLSX, as UTF-8 CSV (BOM for Excel)."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(REGISTER_COLUMNS)
    for c in contracts:
        writer.writerow(["" if v is None else v for v in _register_row(c)])
    return ("﻿" + buf.getvalue()).encode("utf-8")


def contracts_to_register_xlsx(contracts, title: str = "Contract Register") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append(REGISTER_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for c in contracts:
        ws.append(_register_row(c))

    for idx, _ in enumerate(REGISTER_COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = 22

    _append_line_items_sheet(wb, contracts)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


LINE_ITEM_COLUMNS = ["Sr No", "Vendor", "Item", "Unit", "Quantity", "Unit Rate", "Amount"]


def _append_line_items_sheet(wb: Workbook, contracts) -> None:
    """Add a 'Line Items' worksheet flattening every contract's priced rows, so
    unit rates export alongside the register."""
    ws = wb.create_sheet("Line Items")
    ws.append(LINE_ITEM_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for c in contracts:
        vendor = c.vendor.name if c.vendor else c.vendor_name_raw
        for li in (c.line_items or []):
            if not isinstance(li, dict):
                continue
            ws.append([
                c.sr_no,
                vendor,
                li.get("item"),
                li.get("unit"),
                li.get("quantity"),
                li.get("unit_rate"),
                li.get("amount"),
            ])

    for idx, _ in enumerate(LINE_ITEM_COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = 20
