"""Bulk import of existing contracts from an Excel (.xlsx) or CSV file in the
15-column register format (round-trips with the register export).

Parsing and per-row mapping are pure and unit-tested; the DB side (vendor/
department resolution, persistence, dry-run) lives in the API endpoint.
"""
import csv
import io
import re

from .dates import parse_date

# Normalized header -> contract field key. sr_no and contract_link are ignored
# (system-generated).
HEADER_MAP = {
    "signing entity": "signing_entity",
    "vendor": "vendor",
    "vendor name": "vendor",
    "vendor address": "vendor_address",
    "start date": "start_date",
    "end date": "end_date",
    "contract tenure": "contract_tenure",
    "tenure": "contract_tenure",
    "department": "department",
    "po number": "po_number",
    "po": "po_number",
    "contract value": "contract_value",
    "value": "contract_value",
    "currency": "currency",
    "iks signing authority": "iks_signing_authority",
    "vendor signing authority": "vendor_signing_authority",
    "contract service": "contract_service",
    "service": "contract_service",
    "service summary": "service_summary",
    "summary": "service_summary",
    "payment term": "payment_term",
    "payment terms": "payment_term",
    "notice period": "notice_period",
}

DATE_FIELDS = ("start_date", "end_date")


def normalize_header(header: str) -> str:
    return re.sub(r"[\s_]+", " ", (header or "").strip().lower())


def _clean(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def parse_rows(file_bytes: bytes, filename: str) -> list[dict]:
    """Return a list of raw row dicts keyed by contract field, with a 1-based
    `_row` number (accounting for the header row)."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        text = file_bytes.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        table = list(reader)
    elif name.endswith(".xlsx"):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        table = [[c for c in row] for row in ws.iter_rows(values_only=True)]
    else:
        raise ValueError("Unsupported file type — upload a .xlsx or .csv file")

    if not table:
        return []
    header_cells = [normalize_header(str(h)) if h is not None else "" for h in table[0]]
    field_by_col = {i: HEADER_MAP[h] for i, h in enumerate(header_cells) if h in HEADER_MAP}
    if not field_by_col:
        raise ValueError("No recognized register columns found in the header row")

    rows = []
    for idx, raw in enumerate(table[1:], start=2):  # row 1 is the header
        if all(c is None or str(c).strip() == "" for c in raw):
            continue  # skip blank lines
        row = {"_row": idx}
        for col, field in field_by_col.items():
            if col < len(raw):
                row[field] = _clean(raw[col])
        rows.append(row)
    return rows


def parse_value(raw: str | None) -> tuple[float | None, str | None]:
    """Parse a contract value like '1,25,000' / 'INR 50000' / '₹ 1.2' into a number.
    Returns (value, error)."""
    if raw is None or raw == "":
        return None, None
    cleaned = re.sub(r"[^\d.\-]", "", raw.replace(",", ""))
    if cleaned in ("", "-", ".", "-."):
        return None, f"contract value '{raw}' is not a number"
    try:
        return float(cleaned), None
    except ValueError:
        return None, f"contract value '{raw}' is not a number"


def map_row(row: dict) -> tuple[dict, list[str]]:
    """Coerce a raw row into contract fields; returns (fields, errors). Pure — no DB."""
    errors: list[str] = []
    fields: dict = {}

    for key in ("signing_entity", "vendor", "vendor_address", "contract_tenure",
                "department", "po_number", "currency", "iks_signing_authority",
                "vendor_signing_authority", "contract_service", "service_summary",
                "payment_term", "notice_period"):
        if row.get(key):
            fields[key] = row[key]

    for key in DATE_FIELDS:
        raw = row.get(key)
        if raw:
            parsed = parse_date(raw)
            if parsed is None:
                errors.append(f"{key.replace('_', ' ')} '{raw}' is not a valid date")
            else:
                fields[key] = parsed

    value, verr = parse_value(row.get("contract_value"))
    if verr:
        errors.append(verr)
    elif value is not None:
        fields["contract_value"] = value

    return fields, errors
